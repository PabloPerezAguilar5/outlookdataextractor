import os
import win32com.client as win32
import pdfplumber
import camelot
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

def descargar_adjunto_ultimo_correo():
    try:
        outlook = win32.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)  # Carpeta de la bandeja de entrada

        # Buscar el último correo con el asunto que contenga "ADVCP"
        ultimo_correo = None
        for item in reversed(inbox.Items):
            if "ADVCP" in item.Subject:
                ultimo_correo = item
                break

        if ultimo_correo:
            # Descargar el primer archivo adjunto (si es un PDF)
            for attachment in ultimo_correo.Attachments:
                if attachment.FileName.lower().endswith('.pdf'):
                    ruta_descarga = os.path.join(os.path.expanduser("~"), "Desktop", "Nueva carpeta", attachment.FileName)
                    attachment.SaveAsFile(ruta_descarga)
                    print("Archivo adjunto descargado correctamente:", ruta_descarga)
                    return ruta_descarga

        print("No se encontró ningún correo con el asunto que contiene 'ADVCP' o no se encontraron archivos PDF adjuntos.")
        return None
    except Exception as e:
        print("Error al descargar el adjunto del correo:", e)
        return None

def convertir_pdf_a_excel(ruta_pdf, ruta_excel):
    try:
        # Utilizar camelot para extraer datos del PDF y escribirlos en un archivo Excel
        tables = camelot.read_pdf(ruta_pdf, flavor='stream', pages='1-end')
        df = pd.concat([table.df for table in tables], ignore_index=True)

        # Crear un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active

        # Convertir el DataFrame a una hoja de cálculo de Excel
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Guardar el archivo Excel
        wb.save(ruta_excel)

        print("PDF convertido a Excel correctamente.")
        return ruta_excel
    except Exception as e:
        print("Error al convertir el PDF a Excel:", e)
        return None

def extraer_cifra_desde_excel(ruta_archivo_excel):
    try:
        wb = load_workbook(ruta_archivo_excel)
        ws = wb.active
        cifra = ws['G18'].value  # Modificar la celda a leer
        return cifra
    except Exception as e:
        print("Error al extraer la cifra del Excel:", e)
        return None

def enviar_correo(cifra, destinatarios):
    try:
        outlook = win32.Dispatch("Outlook.Application")
        mensaje = outlook.CreateItem(0)
        mensaje.Subject = "Cifra extraída del PDF"
        mensaje.Body = f"La cifra extraída del PDF es: {cifra}"
        mensaje.To = ";".join(destinatarios)  # Concatenar los destinatarios separados por punto y coma
        print("Destinatarios:", mensaje.To)  # Imprimir la lista de destinatarios
        mensaje.Send()
        print("Correo enviado correctamente.")
    except Exception as e:
        print("Error al enviar el correo:", e)

if __name__ == "__main__":
    ruta_archivo_pdf = descargar_adjunto_ultimo_correo()
    if ruta_archivo_pdf:
        ruta_archivo_excel = os.path.splitext(ruta_archivo_pdf)[0] + ".xlsx"
        if convertir_pdf_a_excel(ruta_archivo_pdf, ruta_archivo_excel):
            cifra = extraer_cifra_desde_excel(ruta_archivo_excel)
            if cifra:
                # Correos destinatarios
                destinatarios = ["yourname@yourmail.com", "yourname@yourmail.com"]  

                # Enviar correo con la cifra extraída
                enviar_correo(cifra, destinatarios)
            else:
                print("No se pudo extraer la cifra del archivo Excel.")